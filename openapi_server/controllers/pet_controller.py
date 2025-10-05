import connexion
from typing import Dict
from typing import Tuple
from typing import Union

from openapi_server.models.api_response import ApiResponse  # noqa: E501
from openapi_server.models.error import Error  # noqa: E501
from openapi_server.models.pet import Pet  # noqa: E501
from openapi_server.models.category import Category
from openapi_server.models.tag import Tag
from openapi_server import util

from openpyxl import load_workbook


def add_pet(body):  # noqa: E501
    """Add a new pet to the store.

    Add a new pet to the store. # noqa: E501

    :param pet: Create a new pet in the store
    :type pet: dict | bytes

    :rtype: Union[Pet, Tuple[Pet, int], Tuple[Pet, int, Dict[str, str]]
    """
    pet = body
    if connexion.request.is_json:
        pet = Pet.from_dict(connexion.request.get_json())  # noqa: E501
    return 'do some magic!'


def delete_pet(pet_id, api_key=None):  # noqa: E501
    """Deletes a pet.

    Delete a pet. # noqa: E501

    :param pet_id: Pet id to delete
    :type pet_id: int
    :param api_key: 
    :type api_key: str

    :rtype: Union[None, Tuple[None, int], Tuple[None, int, Dict[str, str]]
    """
    return 'do some magic!'


def find_pets_by_status(status=None):  # noqa: E501
    """Finds Pets by status.

    Multiple status values can be provided with comma separated strings. # noqa: E501

    :param status: Status values that need to be considered for filter
    :type status: str

    :rtype: Union[List[Pet], Tuple[List[Pet], int], Tuple[List[Pet], int, Dict[str, str]]
    """
    # Load the existing workbook
    wb = load_workbook('sample.xlsx', read_only=False, data_only=False)

    # grab the active worksheet
    ws = wb["Pets"]

    # Check for existing ID and find the next empty row
    max_row = 2
    pet_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4] == status:
            print("Find pet with the expected status")
            pet_list.append(row)
        if row[0] is None:
            break
        print(row)
        max_row += 1
    print(f"Max row:{max_row}")

    ws = wb['Categories']
    category_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        category_list.append(row)
        print(row)

    ws = wb['Tags']
    tag_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        tag_list.append(row)
        print(row)

    pets = []
    for pet in pet_list:
        pet_category = None
        category_id = pet[0]
        for category in category_list:
            if category[0] == category_id:
                pet_category = Category(category[0], category[1]).to_dict()
                break
        pet_tags = None
        tags = pet[5]
        if tags is not None:
            tag_id_list = tags.split("-")
            print(f"Find tags {tag_id_list}")
            pet_tags = []
            for tag_id in tag_id_list:
                for tag in tag_list:
                    if tag[0] == int(tag_id):
                        print(f"Find tag with ID {tag_id}")
                        pet_tag = Tag(tag[0], tag[1])
                        pet_tags.append(pet_tag.to_dict())
                        break
        pets.append(Pet(pet[1], pet[2], pet_category, None, pet_tags, status).to_dict())

    # return 'do some magic!'
    return pets


def find_pets_by_tags(tags=None):  # noqa: E501
    """Finds Pets by tags.

    Multiple tags can be provided with comma separated strings. Use tag1, tag2, tag3 for testing. # noqa: E501

    :param tags: Tags to filter by
    :type tags: List[str]

    :rtype: Union[List[Pet], Tuple[List[Pet], int], Tuple[List[Pet], int, Dict[str, str]]
    """
    return 'do some magic!'


def get_pet_by_id(pet_id):  # noqa: E501
    """Find pet by ID.

    Returns a single pet. # noqa: E501

    :param pet_id: ID of pet to return
    :type pet_id: int

    :rtype: Union[Pet, Tuple[Pet, int], Tuple[Pet, int, Dict[str, str]]
    """
    # Load the existing workbook
    wb = load_workbook('sample.xlsx', read_only=False, data_only=False)

    # grab the active worksheet
    ws = wb["Pets"]

    # Check for existing ID and find the next empty row
    max_row = 2
    pet = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == pet_id:
            print("Find pet with the expected ID")
            pet = row
            break
        if row[0] is None:
            break
        print(row)
        max_row += 1
    print(f"Max row:{max_row}")

    if pet is None:
        return (None, 800, Error('800', 'Pet NOt found').to_dict())

    ws = wb['Categories']
    category_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        category_list.append(row)
        print(row)

    ws = wb['Tags']
    tag_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        tag_list.append(row)
        print(row)

    pet_category = None
    category_id = pet[0]
    for category in category_list:
        if category[0] == category_id:
            pet_category = Category(category[0], category[1]).to_dict()
            break
    pet_tags = None
    tags = pet[5]
    if tags is not None:
        tag_id_list = tags.split("-")
        print(f"Find tags {tag_id_list}")
        pet_tags = []
        for tag_id in tag_id_list:
            for tag in tag_list:
                if tag[0] == int(tag_id):
                    print(f"Find tag with ID {tag_id}")
                    pet_tag = Tag(tag[0], tag[1])
                    pet_tags.append(pet_tag.to_dict())
                    break
    return Pet(pet[1], pet[2], pet_category, None, pet_tags, pet[4]).to_dict()

    # return 'do some magic!'
    return pets


def update_pet(body):  # noqa: E501
    """Update an existing pet.

    Update an existing pet by Id. # noqa: E501

    :param pet: Update an existent pet in the store
    :type pet: dict | bytes

    :rtype: Union[Pet, Tuple[Pet, int], Tuple[Pet, int, Dict[str, str]]
    """
    pet = body
    if connexion.request.is_json:
        pet = Pet.from_dict(connexion.request.get_json())  # noqa: E501
    return 'do some magic!'


def update_pet_with_form(pet_id, name=None, status=None):  # noqa: E501
    """Updates a pet in the store with form data.

    Updates a pet resource based on the form data. # noqa: E501

    :param pet_id: ID of pet that needs to be updated
    :type pet_id: int
    :param name: Name of pet that needs to be updated
    :type name: str
    :param status: Status of pet that needs to be updated
    :type status: str

    :rtype: Union[Pet, Tuple[Pet, int], Tuple[Pet, int, Dict[str, str]]
    """
    return 'do some magic!'


def upload_file(pet_id, additional_metadata=None, body=None):  # noqa: E501
    """Uploads an image.

    Upload image of the pet. # noqa: E501

    :param pet_id: ID of pet to update
    :type pet_id: int
    :param additional_metadata: Additional Metadata
    :type additional_metadata: str
    :param body: 
    :type body: str

    :rtype: Union[ApiResponse, Tuple[ApiResponse, int], Tuple[ApiResponse, int, Dict[str, str]]
    """
    body = body
    return 'do some magic!'
