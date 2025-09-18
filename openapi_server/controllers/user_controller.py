import datetime
import connexion
from typing import Dict
from typing import Tuple
from typing import Union

from openapi_server.models.error import Error  # noqa: E501
from openapi_server.models.user import User  # noqa: E501
from openapi_server import util

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.reader.excel import ExcelReader

logged_in_usernames = {}


def create_user(body=None):  # noqa: E501
    """Create user.

    This can only be done by the logged in user. # noqa: E501

    :param user: Created user object
    :type user: dict | bytes

    :rtype: Union[User, Tuple[User, int], Tuple[User, int, Dict[str, str]]
    """
    user = body
    if connexion.request.is_json:
        user = User.from_dict(connexion.request.get_json())  # noqa: E501

    # Load the existing workbook
    wb = load_workbook('sample.xlsx', read_only=False, data_only=False)

    # grab the active worksheet
    ws = wb.active

    # Check for existing ID and find the next empty row
    max_row = 2
    username_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user.id:
            print("ID already exist")
            return (None, 800, Error('800', 'ID already exist').to_dict())
        if row[0] is None:
            break
        print(row)
        username_list.append(row[1])
        max_row += 1
    print(f"Max row:{max_row}")

    if user.username in username_list:
        print("Username already exist")
        return (None, 801, Error('801', 'Username already exist').to_dict())

    # Add user data to the next empty row
    data = [
        user.id, user.username, user.last_name, user.first_name,
        user.email, user.password, user.phone, user.user_status,
        datetime.datetime.now()
    ]
    for col_idx, cell_value in enumerate(data, start=1):
        print(f"Add col_idx:{col_idx}, cell_value:{cell_value}")
        ws.cell(row=max_row, column=col_idx, value=cell_value)

    # Save the file
    wb.save("sample.xlsx")
    # return 'User created successfully'
    return user.to_dict()


def create_users_with_list_input(body=None):  # noqa: E501
    """Creates list of users with given input array.

    Creates list of users with given input array. # noqa: E501

    :param user: 
    :type user: list | bytes

    :rtype: Union[User, Tuple[User, int], Tuple[User, int, Dict[str, str]]
    """
    user = body
    if connexion.request.is_json:
        user = [User.from_dict(d) for d in connexion.request.get_json()]  # noqa: E501
    return 'do some magic!'


def delete_user(username):  # noqa: E501
    """Delete user resource.

    This can only be done by the logged in user. # noqa: E501

    :param username: The name that needs to be deleted
    :type username: str

    :rtype: Union[None, Tuple[None, int], Tuple[None, int, Dict[str, str]]
    """

    # Load the existing workbook
    wb = load_workbook('sample.xlsx', read_only=False, data_only=False)

    # grab the Users worksheet
    ws = wb['Users']

    # Check for existing ID and find the next empty row
    max_row = 2
    username_found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            print("Found user to delete")
            username_found = True
            break
        if row[0] is None:
            break
        print(row)
        max_row += 1
    if not username_found:
        return (None, 404, Error('404', 'User not found').to_dict())
    ws.delete_rows(max_row)
    wb.save("sample.xlsx")
    return None, 200


def get_user_by_name(username):  # noqa: E501
    """Get user by user name.

    Get user detail based on username. # noqa: E501

    :param username: The name that needs to be fetched. Use user1 for testing
    :type username: str

    :rtype: Union[User, Tuple[User, int], Tuple[User, int, Dict[str, str]]
    """

    # Load the existing workbook
    wb = load_workbook('sample.xlsx', read_only=False, data_only=False)

    # grab the Users worksheet
    ws = wb['Users']

    max_row = 2
    user = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            print("Found username")
            user = User(
                id=row[0], username=row[1], last_name=row[2], first_name=row[3],
                email=row[4], password=row[5], phone=row[6], user_status=row[7]
            )
            break
        if row[0] is None:
            break
        print(row)
        max_row += 1
    wb.close()
    if user is None:
        return (None, 404, Error('404', 'User not found').to_dict())
    return user, 200


def login_user(username=None, password=None):  # noqa: E501
    """Logs user into the system.

    Log into the system. # noqa: E501

    :param username: The user name for login
    :type username: str
    :param password: The password for login in clear text
    :type password: str

    :rtype: Union[str, Tuple[str, int], Tuple[str, int, Dict[str, str]]
    """

    # Load the existing workbook
    wb = load_workbook('sample.xlsx', read_only=False, data_only=False)

    # grab the Users worksheet
    ws = wb['Users']

    max_row = 2
    user_can_login = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username and row[5] == password:
            print("User can login")
            logged_in_usernames.update({"username": username})
            user_can_login = True
            break
        if row[0] is None:
            break
        print(row)
        max_row += 1
    wb.close()
    if not user_can_login:
        return (None, 400, Error('400', 'Invalid username/password supplied').to_dict())
    return f'{username} logged in successfully', 200


def logout_user():  # noqa: E501
    """Logs out current logged in user session.

    Log user out of the system. # noqa: E501


    :rtype: Union[None, Tuple[None, int], Tuple[None, int, Dict[str, str]]
    """
    return 'do some magic!'


def update_user(username, body=None):  # noqa: E501
    """Update user resource.

    This can only be done by the logged in user. # noqa: E501

    :param username: name that need to be deleted
    :type username: str
    :param user: Update an existent user in the store
    :type user: dict | bytes

    :rtype: Union[None, Tuple[None, int], Tuple[None, int, Dict[str, str]]
    """
    if username not in logged_in_usernames:
        print(f"{username} not logged in")
        return (None, 400, Error('400', 'User not logged in').to_dict())
    user = body
    if connexion.request.is_json:
        user = User.from_dict(connexion.request.get_json())  # noqa: E501

    # Load the existing workbook
    wb = load_workbook('sample.xlsx', read_only=False, data_only=False)

    # grab the Users worksheet
    ws = wb['Users']

    data = [
        user.id, user.username, user.last_name, user.first_name,
        user.email, user.password, user.phone, user.user_status,
        datetime.datetime.now()
    ]

    max_row = 2
    username_list = []
    username_row = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            print("Found username to update")
            username_row = max_row
        if row[0] is None:
            break
        print(row)
        username_list.append(row[1])
        max_row += 1
    print(f"Max row:{max_row}")
    if username_row is None:
        wb.close()
        return (None, 404, Error('404', 'User not found').to_dict())
    if user.username in username_list:
        wb.close()
        return (None, 801, Error('801', 'Username already exist').to_dict())
    for col_idx, cell_value in enumerate(data, start=1):
        print(f"Update col_idx:{col_idx}, cell_value:{cell_value}")
        ws.cell(row=username_row, column=col_idx, value=cell_value)
    wb.save("sample.xlsx")
    return None, 200
